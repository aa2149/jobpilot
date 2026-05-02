# ============================================================================
# Copyright (c) 2026 [YOUR NAME]. All rights reserved.
# Part of JobPilot — Submitted to the 1000Jobs Final Stage assessment.
# ============================================================================
"""Batch application runner.

Given a list of job URLs and one applicant, runs the agent against each
job sequentially. Returns a structured per-job result + generates a
personalized tracker xlsx that the user can download (or, in production,
have emailed).

Sequential — not parallel — by design:
  - Parallel applications hit the same domain in burst, which is the
    fastest way to get every IP banned.
  - The brief asks for a single agent that handles ONE application
    flawlessly. Batch is ergonomic, not a separate engineering challenge.
  - Per-job pacing also makes the Loom demo legible — you can watch
    each application complete.
"""
from __future__ import annotations

import asyncio
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from jobpilot.api.schemas import Applicant, ApplyOptions
from jobpilot.flow.greenhouse import FlowResult
from jobpilot.flow.router import route_and_apply
from jobpilot.observability.logger import get_run_logger, log_path, make_run_id


@dataclass
class BatchJobResult:
    job_url: str
    run_id: str
    state: str
    fields_filled: int = 0
    questions_answered: int = 0
    resume_uploaded: bool = False
    duration_ms: int = 0
    screenshot: str | None = None
    reason: str | None = None
    company: str = ""
    role: str = ""


@dataclass
class BatchResult:
    batch_id: str
    started_at: str
    finished_at: str
    total_jobs: int
    successes: int
    failures: int
    blocked: int
    results: list[BatchJobResult] = field(default_factory=list)
    tracker_path: str = ""
    email_eml_path: str = ""


async def run_batch(
    *,
    job_urls: list[str],
    applicant: Applicant,
    options: ApplyOptions,
    pause_seconds: float = 8.0,
) -> BatchResult:
    """Run the agent against each URL in sequence. Pause between runs so
    we don't fingerprint as a burst."""
    from datetime import datetime, timezone

    batch_id = f"batch_{datetime.now(timezone.utc).strftime('%Y-%m-%dT%H-%M-%SZ')}"
    started_at = datetime.now(timezone.utc).isoformat()
    results: list[BatchJobResult] = []

    for idx, url in enumerate(job_urls):
        run_id = make_run_id()
        log = get_run_logger(run_id)
        log.info("batch.job.start", batch_id=batch_id, idx=idx, url=url)

        t0 = time.monotonic()
        try:
            flow_result: FlowResult = await route_and_apply(
                run_id=run_id,
                job_url=url,
                applicant=applicant,
                options=options,
            )
        except Exception as e:
            log.exception("batch.job.unhandled", error=str(e))
            results.append(BatchJobResult(
                job_url=url,
                run_id=run_id,
                state="unknown_error",
                reason=f"{type(e).__name__}: {e}",
                duration_ms=int((time.monotonic() - t0) * 1000),
            ))
            continue

        duration = int((time.monotonic() - t0) * 1000)

        # Best-effort company name from the URL
        company, role = _company_role_from_url(url)

        results.append(BatchJobResult(
            job_url=url,
            run_id=run_id,
            state=flow_result.state,
            fields_filled=flow_result.fields_filled,
            questions_answered=flow_result.questions_answered,
            resume_uploaded=flow_result.resume_uploaded,
            duration_ms=duration,
            screenshot=flow_result.screenshot,
            reason=flow_result.reason,
            company=company,
            role=role,
        ))

        # Pause between jobs (skip after the last one)
        if idx < len(job_urls) - 1:
            await asyncio.sleep(pause_seconds)

    finished_at = datetime.now(timezone.utc).isoformat()

    successes = sum(1 for r in results if r.state in ("submitted", "pre_submit_ready"))
    failures = sum(1 for r in results if r.state not in ("submitted", "pre_submit_ready", "blocked_by_bot_detection"))
    blocked = sum(1 for r in results if r.state == "blocked_by_bot_detection")

    # Generate tracker
    tracker_path = generate_batch_tracker(batch_id, applicant, results)

    # "Send" email — stubbed in v1, writes a real .eml file to logs/<batch_id>/email.eml
    from jobpilot.flow.email_stub import send as send_email
    summary = {
        "batch_id": batch_id,
        "total_jobs": len(results),
        "successes": successes,
        "failures": failures,
        "blocked": blocked,
    }
    eml_path = send_email(
        batch_id=batch_id,
        to_addr=applicant.email,
        to_name=f"{applicant.first_name} {applicant.last_name}",
        batch_summary=summary,
        results=results,
        tracker_path=tracker_path,
    )

    return BatchResult(
        batch_id=batch_id,
        started_at=started_at,
        finished_at=finished_at,
        total_jobs=len(job_urls),
        successes=successes,
        failures=failures,
        blocked=blocked,
        results=results,
        tracker_path=str(tracker_path),
        email_eml_path=str(eml_path),
    )


def _company_role_from_url(url: str) -> tuple[str, str]:
    """Best-effort guess from URL pattern."""
    import re
    # job-boards.greenhouse.io/<company>/jobs/<id>
    m = re.search(r"greenhouse\.io/([^/]+)/jobs/", url.lower())
    if m:
        return m.group(1).replace("-", " ").title(), ""
    # company.com/...?gh_jid=<id>
    m = re.search(r"https?://(?:www\.)?([^./]+)\.", url.lower())
    if m:
        return m.group(1).title(), ""
    return "Unknown", ""


def generate_batch_tracker(
    batch_id: str,
    applicant: Applicant,
    results: list[BatchJobResult],
) -> Path:
    """Create a personalized tracker xlsx for this batch.

    This is what would be emailed to the applicant in production. For the
    demo, the file is placed in logs/<batch_id>/tracker.xlsx and the
    response includes the path so the frontend can offer a download.
    """
    from datetime import date
    from jobpilot.observability.logger import run_dir
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

    THIN = Side(border_style="thin", color="CCCCCC")
    ALL_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    HEADER_FILL = PatternFill("solid", fgColor="161510")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=11, color="161510")

    out_dir = run_dir(batch_id)
    out_path = out_dir / "tracker.xlsx"

    wb = openpyxl.Workbook()
    wb.properties.creator = applicant.first_name + " " + applicant.last_name
    wb.properties.title = f"JobPilot tracker — {batch_id}"

    ws = wb.active
    ws.title = "Tracker"

    # Title
    ws["A1"] = f"Application Tracker — {applicant.first_name} {applicant.last_name}"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="161510")
    ws["A2"] = (
        f"Auto-generated by JobPilot on {date.today().isoformat()} · "
        f"Batch {batch_id} · {len(results)} applications · "
        "Update statuses below as you hear back."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws.row_dimensions[1].height = 26

    headers = [
        "Date Applied", "Company", "Role", "Job URL", "State",
        "Fields Filled", "Questions Answered", "Resume Uploaded",
        "Status", "Response", "Round 1", "Round 2", "Round 3",
        "Last Update", "Notes", "Run ID",
    ]
    col_widths = [12, 22, 24, 38, 18, 12, 18, 14, 16, 18, 12, 12, 12, 12, 36, 28]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28

    today = date.today().isoformat()
    for r_idx, r in enumerate(results, start=5):
        applied = r.state in ("submitted", "pre_submit_ready")
        initial_status = "Applied" if r.state == "submitted" else \
                         "Pre-Submit (review)" if r.state == "pre_submit_ready" else \
                         "Failed"
        row = [
            today, r.company, r.role, r.job_url, r.state,
            r.fields_filled, r.questions_answered, "yes" if r.resume_uploaded else "no",
            initial_status, "Awaiting Response", "", "", "",
            today, r.reason or "", r.run_id,
        ]
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = ALL_THIN

    # Reserve 30 empty rows
    for r in range(5 + len(results), 5 + len(results) + 30):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=c_idx).border = ALL_THIN

    # Status dropdown (col 9 = I)
    last = 4 + len(results) + 30
    statuses = ["Applied", "In Review", "Phone Screen", "Round 1", "Round 2", "Round 3",
                "Take-Home", "Offer", "Negotiating", "Accepted", "Rejected",
                "Ghosted", "Withdrew", "Pre-Submit (review)", "Failed"]
    sdv = DataValidation(type="list", formula1='"' + ",".join(statuses) + '"', allow_blank=True)
    sdv.add(f"I5:I{last}")
    ws.add_data_validation(sdv)

    # Response dropdown (col 10 = J)
    rdv = DataValidation(
        type="list",
        formula1='"Awaiting Response,Responded,Auto-Rejected,No Response (>21d)"',
        allow_blank=True,
    )
    rdv.add(f"J5:J{last}")
    ws.add_data_validation(rdv)

    # Round dropdowns (K, L, M)
    round_dv = DataValidation(
        type="list",
        formula1='"Scheduled,Done-Pass,Done-Fail,Skipped,N/A"',
        allow_blank=True,
    )
    round_dv.add(f"K5:M{last}")
    ws.add_data_validation(round_dv)

    # Status conditional formatting
    color_map = {
        "Applied": ("D97757", "FCEEE6"),
        "Round 1": ("92400E", "FEF3C7"),
        "Round 2": ("92400E", "FEF3C7"),
        "Round 3": ("92400E", "FEF3C7"),
        "Offer": ("1F4D3F", "E1ECE6"),
        "Accepted": ("1F4D3F", "E1ECE6"),
        "Rejected": ("9B2C2C", "FCE7E7"),
        "Ghosted": ("9B2C2C", "FCE7E7"),
        "Failed": ("9B2C2C", "FCE7E7"),
    }
    for status, (fg, bg) in color_map.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{status}"'],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name="Calibri", size=11, bold=True, color=fg),
        )
        ws.conditional_formatting.add(f"I5:I{last}", rule)

    # Round outcome formatting
    pass_rule = CellIsRule(
        operator="equal", formula=['"Done-Pass"'],
        fill=PatternFill("solid", fgColor="E1ECE6"),
        font=Font(name="Calibri", size=11, bold=True, color="1F4D3F"),
    )
    fail_rule = CellIsRule(
        operator="equal", formula=['"Done-Fail"'],
        fill=PatternFill("solid", fgColor="FCE7E7"),
        font=Font(name="Calibri", size=11, bold=True, color="9B2C2C"),
    )
    for rule in (pass_rule, fail_rule):
        ws.conditional_formatting.add(f"K5:M{last}", rule)

    ws.freeze_panes = "A5"

    # Summary sheet with the ask: where the agent applied
    summary = wb.create_sheet("Summary", 0)
    summary["A1"] = f"Hi {applicant.first_name} — here's where JobPilot applied for you."
    summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="161510")
    summary["A2"] = (
        f"Generated {today} · {len(results)} applications attempted · "
        f"{sum(1 for r in results if r.state == 'submitted')} submitted · "
        f"{sum(1 for r in results if r.state == 'pre_submit_ready')} review-ready · "
        f"{sum(1 for r in results if r.state == 'blocked_by_bot_detection')} blocked"
    )
    summary["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")

    summary["A4"] = "Now it's your turn:"
    summary["A4"].font = Font(name="Calibri", size=11, bold=True, color="D97757")
    tips = [
        "1. Open the 'Tracker' sheet (tab below)",
        "2. As you hear back from each company, set the Status column",
        "3. Track each interview round in Round 1 / Round 2 / Round 3 columns",
        "4. After 21 days with no response, mark Status as 'Ghosted' — it's data, not silence",
        "5. The Notes column is for anything: recruiter name, salary range, gut feel",
        "",
        "When you've heard from ~10 companies, you'll have real signal about your funnel:",
        "  - If application→response rate < 10%: your resume or fit-targeting needs work",
        "  - If response→Round 1 rate < 30%: rethink your phone-screen prep",
        "  - If Round 1→Round 2 rate < 50%: practice interview answers (try the Story Bank)",
        "",
        "JobPilot only helps with the volume problem. You drive the rest.",
    ]
    for i, t in enumerate(tips, start=5):
        summary.cell(row=i, column=1, value=t).font = Font(name="Calibri", size=11, color="161510")
    summary.column_dimensions["A"].width = 90

    wb.save(out_path)
    return out_path
