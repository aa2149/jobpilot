# ============================================================================
# Copyright (c) 2026 Areej Ahmed. All rights reserved.
# Part of JobPilot — Submitted to the 1000Jobs Final Stage assessment.
# Licensed under the JobPilot Evaluation & Personal-Use License.
# See LICENSE and NOTICE.md in the repository root.
# ============================================================================
"""Build the JobPilot application tracker.

A real, working Excel workbook with:
  - Tracker sheet: one row per application, with statuses, dates, links
  - Interview Rounds sheet: separate detailed log of each interview
  - Story Bank sheet: STAR+R stories accumulated across applications
  - Dashboard sheet: live counts via COUNTIF formulas
  - Status Legend sheet: canonical statuses + color codes
  - Platforms sheet: the platform graph as reference

This file is what a candidate actually uses week-to-week. The agent
appends rows to Tracker; the candidate fills in interview details by hand.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# Style tokens (consistent across sheets)
# =============================================================================

INK = "161510"
INK_SOFT = "4A4940"
INK_LIGHT = "B8B5A6"
ACCENT = "D97757"
ACCENT_SOFT = "FCEEE6"
EMERALD = "1F4D3F"
EMERALD_SOFT = "E1ECE6"
ROSE = "9B2C2C"
ROSE_SOFT = "FCE7E7"
AMBER = "92400E"
AMBER_SOFT = "FEF3C7"
PAPER = "FAFAF7"
THIN = Side(border_style="thin", color="CCCCCC")
ALL_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

HEADER_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11, color=INK)
SMALL_FONT = Font(name="Calibri", size=10, color=INK_SOFT)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=INK)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color=INK_SOFT)


# =============================================================================
# Canonical statuses
# =============================================================================

STATUSES = [
    ("Saved",              "Job saved but not yet applied",                                 INK_LIGHT,    PAPER),
    ("Applied",            "Application submitted",                                          ACCENT,       ACCENT_SOFT),
    ("In Review",          "Application acknowledged or under review",                      AMBER,        AMBER_SOFT),
    ("Phone Screen",       "Recruiter / HR initial call scheduled or done",                 AMBER,        AMBER_SOFT),
    ("Interview R1",       "First-round interview (technical or hiring manager)",           AMBER,        AMBER_SOFT),
    ("Interview R2",       "Second-round interview (panel, deep dive, case)",               AMBER,        AMBER_SOFT),
    ("Interview R3",       "Third-round / final / leadership / executive",                  AMBER,        AMBER_SOFT),
    ("Take-Home",          "Take-home assignment in progress",                              AMBER,        AMBER_SOFT),
    ("Offer",              "Offer extended",                                                 EMERALD,      EMERALD_SOFT),
    ("Negotiating",        "Offer received, negotiating terms",                              EMERALD,      EMERALD_SOFT),
    ("Accepted",           "Offer accepted",                                                 EMERALD,      EMERALD_SOFT),
    ("Rejected",           "Rejected by company",                                            ROSE,         ROSE_SOFT),
    ("Withdrew",           "Candidate withdrew",                                             INK_SOFT,     PAPER),
    ("Ghosted",            "No response after 3+ weeks (treat as soft reject)",              ROSE,         ROSE_SOFT),
    ("Closed",             "Posting closed without outcome",                                 INK_SOFT,     PAPER),
]

RESPONSE_STATES = ["Awaiting Response", "Responded", "Auto-Rejected", "No Response (>21d)"]
WORK_MODES_LIST = ["onsite", "hybrid", "remote", "freelance"]


# =============================================================================
# Sheet builders
# =============================================================================

def _set_header(ws, headers, row=1, col_widths=None):
    """Write a styled header row and set column widths."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def _add_title(ws, title, subtitle=""):
    """Add a title row at the top, leaving room above the data."""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 28


def build_tracker_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Tracker"
    _add_title(ws,
               "JobPilot — Application Tracker",
               "© 2026 Areej Ahmed · All rights reserved · Submitted to 1000Jobs Final Stage for evaluation only · See LICENSE")

    headers = [
        "ID", "Date Saved", "Date Applied", "Company", "Role", "Work Mode",
        "Platform", "Job URL", "Location", "Salary Range",
        "Fit Score", "Archetype Match", "Resume Version",
        "Status", "Response", "Last Update", "Days Since",
        "Recruiter / Contact", "Recruiter Email/LinkedIn",
        "Notes", "Run ID (logs)",
    ]
    col_widths = [
        6, 12, 12, 22, 28, 11,
        14, 32, 16, 14,
        8, 16, 14,
        15, 18, 12, 11,
        22, 28, 38, 22,
    ]
    _set_header(ws, headers, row=4, col_widths=col_widths)

    # Sample rows so the user sees what good data looks like
    samples = [
        [1, "2026-04-28", "2026-04-29", "Grammarly", "Senior Product Manager", "remote",
         "Greenhouse", "https://job-boards.greenhouse.io/grammarly/jobs/7767680", "Remote · US",
         "$180-220k", 0.82, "women_forward + stable_provider", "v3-pm-grammarly",
         "Phone Screen", "Responded", "2026-05-01", None,
         "Sarah Chen", "sarah.chen@grammarly.com", "Recruiter very responsive. Mentioned WLB explicitly. Phone screen Mon 11am.",
         "run_2026-04-29T14-12-22Z_8f3a"],
        [2, "2026-04-25", "2026-04-26", "Modern Health", "Senior PMM", "remote",
         "Greenhouse", "https://job-boards.greenhouse.io/modernhealth/jobs/8465250002", "Remote · US",
         "$160-190k", 0.78, "women_forward + mission_believer", "v2-pmm-mh",
         "Interview R1", "Responded", "2026-04-30", None,
         "Maya Patel", "maya@modernhealth.com", "Hiring manager loved the case study. R2 scheduled.",
         "run_2026-04-26T09-44-08Z_b21c"],
        [3, "2026-04-30", "2026-04-30", "Geotab", "Senior PM, Go App", "remote",
         "Greenhouse", "https://job-boards.greenhouse.io/geotab/jobs/5041684008", "Remote · CA",
         "$150-180k", 0.61, "stable_provider", "v3-pm-geotab",
         "Applied", "Awaiting Response", "2026-04-30", None,
         "", "", "Auto-applied via JobPilot. Below 0.7 archetype match — flagged.",
         "run_2026-04-30T16-02-44Z_c98d"],
        [4, "2026-04-22", "", "Anthropic", "Member of Technical Staff", "hybrid",
         "Custom (anthropic.com/careers)", "https://www.anthropic.com/careers", "San Francisco",
         "$240-300k", 0.55, "deep_specialist", "v4-eng-anthropic",
         "Saved", "—", "2026-04-22", None,
         "", "", "Saved — need to network in before applying. Noted 2 SF-based connections on LinkedIn.",
         ""],
    ]
    for r_idx, row in enumerate(samples, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = ALL_THIN

    # "Days Since" formula — column Q (17), references P (16) Last Update
    for r in range(5, 5 + len(samples)):
        ws.cell(row=r, column=17, value=f"=IF(P{r}=\"\",\"\",TODAY()-P{r})")

    # Reserve empty rows for new entries (50 more)
    for r in range(5 + len(samples), 5 + len(samples) + 50):
        ws.cell(row=r, column=17, value=f"=IF(P{r}=\"\",\"\",TODAY()-P{r})")
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=c_idx).border = ALL_THIN

    # Data validation: Status column (N = 14)
    status_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(s[0] for s in STATUSES)}"',
        allow_blank=True,
    )
    status_dv.add(f"N5:N{4 + len(samples) + 50}")
    ws.add_data_validation(status_dv)

    # Data validation: Response column (O = 15)
    resp_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(RESPONSE_STATES)}"',
        allow_blank=True,
    )
    resp_dv.add(f"O5:O{4 + len(samples) + 50}")
    ws.add_data_validation(resp_dv)

    # Data validation: Work Mode column (F = 6)
    mode_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(WORK_MODES_LIST)}"',
        allow_blank=True,
    )
    mode_dv.add(f"F5:F{4 + len(samples) + 50}")
    ws.add_data_validation(mode_dv)

    # Conditional formatting on Status column (N) — color by status
    for status, _, fg, bg in STATUSES:
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{status}"'],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name="Calibri", size=11, bold=True, color=fg),
        )
        ws.conditional_formatting.add(f"N5:N{4 + len(samples) + 50}", rule)

    # Highlight "Days Since" > 21 in soft rose (a stale application)
    stale_rule = CellIsRule(
        operator="greaterThan",
        formula=["21"],
        fill=PatternFill("solid", fgColor=ROSE_SOFT),
        font=Font(name="Calibri", size=11, color=ROSE),
    )
    ws.conditional_formatting.add(f"Q5:Q{4 + len(samples) + 50}", stale_rule)

    # Freeze the header row
    ws.freeze_panes = "A5"


def build_interview_rounds_sheet(wb: Workbook):
    ws = wb.create_sheet("Interview Rounds")
    _add_title(ws,
               "Interview Rounds — Detail Log",
               "One row per interview round per application. Cross-reference Application ID with the Tracker sheet.")

    headers = [
        "App ID", "Company", "Role", "Round", "Round Type",
        "Date", "Time", "Duration (min)",
        "Interviewer Name", "Interviewer Title", "Interviewer LinkedIn",
        "Format", "Outcome", "Confidence",
        "Questions Asked", "What I Did Well", "What I'd Improve", "Stories Used",
        "Follow-Up Sent?", "Thank-You Sent?", "Notes",
    ]
    col_widths = [
        6, 18, 24, 10, 20,
        12, 8, 12,
        20, 24, 28,
        14, 14, 12,
        38, 30, 30, 30,
        14, 14, 38,
    ]
    _set_header(ws, headers, row=4, col_widths=col_widths)

    samples = [
        [1, "Grammarly", "Senior PM", "R1", "Phone Screen — Recruiter",
         "2026-05-04", "11:00", 30,
         "Sarah Chen", "Senior Recruiter", "linkedin.com/in/sarahchen-grammarly",
         "Phone", "Pass", "High",
         "Why Grammarly? What's your PM philosophy? Why remote?",
         "Articulated WLB priorities clearly. Story about Q3 launch landed.",
         "Could have asked more about team structure.",
         "Q3 launch (story-1), team-restructuring (story-3)",
         "Yes", "Yes", "Sarah is moving me to HM round. Said comp range was confirmed."],
        [2, "Modern Health", "Senior PMM", "R1", "Hiring Manager",
         "2026-05-02", "14:00", 60,
         "Maya Patel", "Director of Product Marketing", "linkedin.com/in/mayapatel",
         "Video", "Pass", "High",
         "Walk me through a launch. How do you think about positioning? Mental health POV?",
         "Strong launch case study. Genuine answer on mental health.",
         "Got tripped up on segmentation question. Practice this.",
         "Q3 launch (story-1), product-positioning (story-2)",
         "Yes", "Yes", "Maya great — felt aligned. R2 with VP next week. Asked for case study brief."],
        [2, "Modern Health", "Senior PMM", "R2", "VP / Skip-Level",
         "2026-05-09", "11:00", 45,
         "Jordan Lee", "VP Marketing", "linkedin.com/in/jordanlee",
         "Video", "Pending", "Medium",
         "Long-term GTM vision. Failure story. Why us over Calm?",
         "Failure story landed. Gave honest answer on competitor analysis.",
         "Could have prepared a stronger 'why us over Calm'.",
         "Q3-launch-failed-experiment (story-4)",
         "", "", "Awaiting feedback. Maya said 'positive signals'."],
    ]
    for r_idx, row in enumerate(samples, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = ALL_THIN
        ws.row_dimensions[r_idx].height = 78

    # Reserve empty rows
    for r in range(5 + len(samples), 5 + len(samples) + 30):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=c_idx).border = ALL_THIN

    # DV: Round
    round_dv = DataValidation(type="list",
                              formula1='"R1,R2,R3,R4,Take-Home,Final,Reference Check"',
                              allow_blank=True)
    round_dv.add(f"D5:D{4 + len(samples) + 30}")
    ws.add_data_validation(round_dv)

    # DV: Outcome
    outcome_dv = DataValidation(type="list",
                                formula1='"Pass,Fail,Pending,Withdrew"',
                                allow_blank=True)
    outcome_dv.add(f"M5:M{4 + len(samples) + 30}")
    ws.add_data_validation(outcome_dv)

    # DV: Confidence
    conf_dv = DataValidation(type="list",
                             formula1='"High,Medium,Low"',
                             allow_blank=True)
    conf_dv.add(f"N5:N{4 + len(samples) + 30}")
    ws.add_data_validation(conf_dv)

    # DV: Format
    fmt_dv = DataValidation(type="list",
                            formula1='"Phone,Video,In-Person,Async (recorded),Take-Home"',
                            allow_blank=True)
    fmt_dv.add(f"L5:L{4 + len(samples) + 30}")
    ws.add_data_validation(fmt_dv)

    # DV: Yes/No for follow-ups (S, T)
    yn_dv = DataValidation(type="list",
                           formula1='"Yes,No,N/A"',
                           allow_blank=True)
    yn_dv.add(f"S5:T{4 + len(samples) + 30}")
    ws.add_data_validation(yn_dv)

    # Conditional formatting on Outcome
    pass_rule = CellIsRule(operator="equal", formula=['"Pass"'],
                           fill=PatternFill("solid", fgColor=EMERALD_SOFT),
                           font=Font(name="Calibri", size=11, bold=True, color=EMERALD))
    fail_rule = CellIsRule(operator="equal", formula=['"Fail"'],
                           fill=PatternFill("solid", fgColor=ROSE_SOFT),
                           font=Font(name="Calibri", size=11, bold=True, color=ROSE))
    pending_rule = CellIsRule(operator="equal", formula=['"Pending"'],
                              fill=PatternFill("solid", fgColor=AMBER_SOFT),
                              font=Font(name="Calibri", size=11, color=AMBER))
    for rule in (pass_rule, fail_rule, pending_rule):
        ws.conditional_formatting.add(f"M5:M{4 + len(samples) + 30}", rule)

    ws.freeze_panes = "A5"


def build_story_bank_sheet(wb: Workbook):
    ws = wb.create_sheet("Story Bank")
    _add_title(ws,
               "Story Bank — STAR + R",
               "Reusable behavioral interview stories. Inspired by career-ops by santifer. Aim for 5-10 strong stories that can be reshaped to answer most behavioral questions.")

    headers = [
        "Story ID", "Title", "Theme",
        "Situation", "Task", "Action", "Result", "Reflection",
        "Used In", "Times Used", "Strength (1-5)", "Refresh Date",
    ]
    col_widths = [10, 28, 16, 36, 30, 38, 28, 28, 26, 11, 13, 13]
    _set_header(ws, headers, row=4, col_widths=col_widths)

    sample_stories = [
        ["story-1", "Q3 launch — pivoting under deadline", "Leadership / Adaptability",
         "Q3 2024, 4 weeks before our flagship launch, the eng team flagged that the ML inference cost would 5x in production vs staging.",
         "I owned the launch. Decision: cut features to hit cost target, or push date and risk competitor.",
         "Spent a day with eng to map cost vs feature impact. Cut two features that drove only 12% of value but 60% of cost. Communicated trade-off to exec team and to design. Negotiated with CFO for a soft cost ceiling rather than hard.",
         "Launched on time at 70% of feature scope. Cost came in 18% under target. Customer feedback was positive on launched features; the cut features later returned in v1.1.",
         "What I'd do differently: surface cost-of-inference earlier in the spec process. Now we run a 'cost diligence' check at design-review.",
         "Grammarly R1, Modern Health R1", 2, 5, "2026-04-28"],
        ["story-2", "Product positioning vs incumbent", "Strategy / PMM",
         "Joined a team launching a category-creator product into a mature space dominated by an incumbent with 8x our budget.",
         "Define positioning that wouldn't get steamrolled.",
         "Researched 200 customer reviews of the incumbent — clustered into 3 unmet-need themes. Built positioning around the strongest theme ('built for teams of 10-50, not 1000+'). Validated with 12 customer calls before locking.",
         "Won 4 of 5 bake-offs vs incumbent in pilot quarter. Positioning held through GTM expansion.",
         "Customer reviews are an underused source of strategic insight. Now I do this on every new product I touch.",
         "Modern Health R1", 1, 4, "2026-04-28"],
        ["story-3", "Restructuring a stuck team", "Leadership / People",
         "Inherited a team of 6 PMs with low morale, last to ship a feature 7 months earlier.",
         "Get them shipping again without an offsite or the 'inspirational speech' approach.",
         "1:1s with each PM in week 1: pure listening, no advice. Found 4 of 6 had unclear scope. Re-cut scopes with each, made trade-offs explicit. Killed two stale projects publicly.",
         "First ship in week 5. By month 3, team velocity 2.5x. One PM left (good outcome — not aligned with new direction).",
         "Underrated lever: just making scope explicit. Most stuck teams aren't unmotivated, they're confused.",
         "Grammarly R1", 1, 5, "2026-04-28"],
        ["story-4", "Failed experiment that I had to defend", "Failure / Reflection",
         "Ran an A/B test that I was sure would lift activation 10%. Designed, shipped, ran for 3 weeks. Result: -2% activation.",
         "Communicate the failure up the chain without burying it.",
         "Wrote a post-mortem the day results came in. Admitted my prior was wrong. Identified 2 confounders I should have caught upfront. Shared with the broader product org — not just my manager.",
         "Manager and skip-level both told me later they remembered the post-mortem more than my wins that quarter. Got promoted 6 months later partly on the strength of how I handled it.",
         "The willingness to publish your own failure publicly is a force multiplier. People remember it.",
         "Modern Health R2", 1, 4, "2026-04-28"],
    ]
    for r_idx, row in enumerate(sample_stories, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = ALL_THIN
        ws.row_dimensions[r_idx].height = 110

    # Reserve empty rows for new stories
    for r in range(5 + len(sample_stories), 5 + len(sample_stories) + 8):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=c_idx).border = ALL_THIN

    # DV: Strength
    strength_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    strength_dv.add(f"K5:K{4 + len(sample_stories) + 8}")
    ws.add_data_validation(strength_dv)

    ws.freeze_panes = "A5"


def build_dashboard_sheet(wb: Workbook):
    ws = wb.create_sheet("Dashboard", 0)  # insert at position 0 = first
    _add_title(ws,
               "JobPilot — Dashboard",
               "© 2026 Areej Ahmed · All rights reserved · 1000Jobs Final Stage submission")

    # Title row spans columns
    for col_letter in ("A", "B", "C", "D"):
        ws.column_dimensions[col_letter].width = 22

    # Section: Counts by status
    ws["A4"] = "BY STATUS"
    ws["A4"].font = Font(name="Calibri", size=10, bold=True, color=ACCENT)

    status_rows = [
        ("Total saved",           '=COUNTIF(Tracker!N:N,"Saved")'),
        ("Total applied",         '=COUNTIF(Tracker!N:N,"Applied")+COUNTIF(Tracker!N:N,"In Review")+COUNTIF(Tracker!N:N,"Phone Screen")+COUNTIF(Tracker!N:N,"Interview R1")+COUNTIF(Tracker!N:N,"Interview R2")+COUNTIF(Tracker!N:N,"Interview R3")+COUNTIF(Tracker!N:N,"Take-Home")+COUNTIF(Tracker!N:N,"Offer")+COUNTIF(Tracker!N:N,"Negotiating")+COUNTIF(Tracker!N:N,"Accepted")+COUNTIF(Tracker!N:N,"Rejected")+COUNTIF(Tracker!N:N,"Ghosted")+COUNTIF(Tracker!N:N,"Closed")'),
        ("In active interview",   '=COUNTIF(Tracker!N:N,"Phone Screen")+COUNTIF(Tracker!N:N,"Interview R1")+COUNTIF(Tracker!N:N,"Interview R2")+COUNTIF(Tracker!N:N,"Interview R3")+COUNTIF(Tracker!N:N,"Take-Home")'),
        ("Offers received",       '=COUNTIF(Tracker!N:N,"Offer")+COUNTIF(Tracker!N:N,"Negotiating")+COUNTIF(Tracker!N:N,"Accepted")'),
        ("Rejected / ghosted",    '=COUNTIF(Tracker!N:N,"Rejected")+COUNTIF(Tracker!N:N,"Ghosted")'),
        ("Awaiting response",     '=COUNTIF(Tracker!O:O,"Awaiting Response")'),
        ("Stale (>21 days)",      '=COUNTIF(Tracker!Q:Q,">21")'),
    ]
    for i, (label, formula) in enumerate(status_rows, start=5):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        ws.cell(row=i, column=1).border = ALL_THIN
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=2).font = Font(name="Calibri", size=12, bold=True, color=INK)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=2).border = ALL_THIN

    # Section: Funnel rate
    fr = 5 + len(status_rows) + 2
    ws.cell(row=fr - 1, column=1, value="FUNNEL").font = Font(name="Calibri", size=10, bold=True, color=ACCENT)

    funnel_rows = [
        ("Application → Interview rate",
         '=IFERROR((COUNTIF(Tracker!N:N,"Phone Screen")+COUNTIF(Tracker!N:N,"Interview R1")+COUNTIF(Tracker!N:N,"Interview R2")+COUNTIF(Tracker!N:N,"Interview R3")+COUNTIF(Tracker!N:N,"Offer")+COUNTIF(Tracker!N:N,"Negotiating")+COUNTIF(Tracker!N:N,"Accepted"))/B6,0)',
         "0.0%"),
        ("Interview → Offer rate",
         '=IFERROR((COUNTIF(Tracker!N:N,"Offer")+COUNTIF(Tracker!N:N,"Negotiating")+COUNTIF(Tracker!N:N,"Accepted"))/(COUNTIF(Tracker!N:N,"Phone Screen")+COUNTIF(Tracker!N:N,"Interview R1")+COUNTIF(Tracker!N:N,"Interview R2")+COUNTIF(Tracker!N:N,"Interview R3")+COUNTIF(Tracker!N:N,"Offer")+COUNTIF(Tracker!N:N,"Negotiating")+COUNTIF(Tracker!N:N,"Accepted")),0)',
         "0.0%"),
        ("Average fit score (applied)",
         '=IFERROR(AVERAGEIFS(Tracker!K:K,Tracker!C:C,"<>"),0)',
         "0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(funnel_rows, start=fr):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        ws.cell(row=i, column=1).border = ALL_THIN
        c = ws.cell(row=i, column=2, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color=INK)
        c.alignment = Alignment(horizontal="right")
        c.border = ALL_THIN
        c.number_format = fmt

    # Section: By work mode
    wm_start = fr + len(funnel_rows) + 2
    ws.cell(row=wm_start - 1, column=1, value="BY WORK MODE").font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
    for i, mode in enumerate(WORK_MODES_LIST, start=wm_start):
        ws.cell(row=i, column=1, value=mode.capitalize()).font = BODY_FONT
        ws.cell(row=i, column=1).border = ALL_THIN
        c = ws.cell(row=i, column=2, value=f'=COUNTIF(Tracker!F:F,"{mode}")')
        c.font = Font(name="Calibri", size=12, bold=True, color=INK)
        c.alignment = Alignment(horizontal="right")
        c.border = ALL_THIN

    # Right column: Tips / read-me
    ws.cell(row=4, column=4, value="HOW TO USE").font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
    tips = [
        "1. The agent appends new applications to the Tracker sheet.",
        "2. Update Status as you progress — drop-downs are pre-filled.",
        "3. For each interview, add a row to the Interview Rounds sheet.",
        "4. Build your Story Bank as you interview. Aim for 5-10 strong stories.",
        "5. The dashboard updates live from formulas.",
        "",
        "STALE APPLICATIONS",
        "Any row with Days Since > 21 in the Tracker is auto-highlighted.",
        "Either follow up, or move it to Ghosted.",
        "",
        "STORY BANK STRATEGY",
        "Review the Story Bank before each interview — pick 3-4 stories that fit",
        "the company's likely questions. After the interview, log which stories you used",
        "in the Interview Rounds sheet, then update Times Used in the Story Bank.",
    ]
    for i, tip in enumerate(tips, start=5):
        c = ws.cell(row=i, column=4, value=tip)
        if tip.isupper():
            c.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
        else:
            c.font = SMALL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["D"].width = 60


def build_status_legend_sheet(wb: Workbook):
    ws = wb.create_sheet("Status Legend")
    _add_title(ws, "Status Legend",
               "Canonical statuses used by the Tracker dropdown. Color-coded.")

    headers = ["Status", "Description", "Color (text/bg)"]
    _set_header(ws, headers, row=4, col_widths=[18, 60, 22])

    for i, (status, desc, fg, bg) in enumerate(STATUSES, start=5):
        c1 = ws.cell(row=i, column=1, value=status)
        c1.font = Font(name="Calibri", size=11, bold=True, color=fg)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.border = ALL_THIN
        c1.alignment = Alignment(vertical="center")

        c2 = ws.cell(row=i, column=2, value=desc)
        c2.font = BODY_FONT
        c2.border = ALL_THIN
        c2.alignment = Alignment(vertical="center", wrap_text=True)

        c3 = ws.cell(row=i, column=3, value=f"#{fg} on #{bg}")
        c3.font = Font(name="Consolas", size=10, color=INK_SOFT)
        c3.border = ALL_THIN
        c3.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 22


def build_platforms_sheet(wb: Workbook):
    ws = wb.create_sheet("Platforms")
    _add_title(ws, "Platform Reference",
               "Master list of supported and planned platforms. Filter by work mode + region.")

    headers = ["Platform", "Domain", "Kind", "Work Modes", "Regions", "Status", "Notes"]
    _set_header(ws, headers, row=4, col_widths=[24, 28, 14, 24, 18, 11, 60])

    # Import here to avoid circular imports if this module is imported from anywhere weird
    from jobpilot.platforms.graph import PLATFORMS

    rows = []
    for p in PLATFORMS.values():
        rows.append([
            p.display_name,
            p.domain,
            p.kind,
            ", ".join(p.work_modes),
            ", ".join(p.regions),
            p.status,
            p.notes,
        ])
    rows.sort(key=lambda r: (r[5] != "live", r[2], r[0]))  # live first, then by kind, then name

    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = ALL_THIN
        # Status color
        status = row[5]
        status_cell = ws.cell(row=r_idx, column=6)
        if status == "live":
            status_cell.fill = PatternFill("solid", fgColor=EMERALD_SOFT)
            status_cell.font = Font(name="Calibri", size=11, bold=True, color=EMERALD)
        elif status == "stub":
            status_cell.fill = PatternFill("solid", fgColor=AMBER_SOFT)
            status_cell.font = Font(name="Calibri", size=11, color=AMBER)
        else:
            status_cell.fill = PatternFill("solid", fgColor=PAPER)
            status_cell.font = Font(name="Calibri", size=11, color=INK_LIGHT)

    ws.freeze_panes = "A5"


# =============================================================================
# Main
# =============================================================================

def build(output_path: str) -> None:
    wb = Workbook()

    # Document properties — these show up in File → Properties / xlsx metadata
    wb.properties.creator = "Areej Ahmed"
    wb.properties.title = "JobPilot — Application Tracker"
    wb.properties.subject = "1000Jobs Final Stage submission"
    wb.properties.description = (
        "© 2026 Areej Ahmed. All rights reserved. Original work submitted to "
        "the 1000Jobs Final Stage assessment for evaluation only. "
        "See LICENSE and NOTICE.md in the source repository."
    )
    wb.properties.keywords = "JobPilot, application tracker, 1000Jobs, copyright reserved"
    wb.properties.lastModifiedBy = "Areej Ahmed"

    build_tracker_sheet(wb)             # this is the active sheet (Sheet1 → Tracker)
    build_dashboard_sheet(wb)           # inserted at index 0
    build_interview_rounds_sheet(wb)
    build_story_bank_sheet(wb)
    build_status_legend_sheet(wb)
    build_platforms_sheet(wb)
    wb.save(output_path)


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "jobpilot_tracker.xlsx"
    build(out)
    print(f"Wrote: {out}")
